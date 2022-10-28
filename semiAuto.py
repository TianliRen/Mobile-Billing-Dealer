from PyQt5.QtWidgets import QDialog, QMessageBox
from PyQt5 import QtWidgets, uic
from PyQt5.QtGui import QIcon, QFont
from openpyxl import *
import sys
import ctypes
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("myappid")


class BillWin(QDialog):
    def __init__(self):
        super(BillWin, self).__init__()
        uic.loadUi('semiAuto.ui', self)

        # Decorate the window with some attributes
        self.setWindowIcon(QIcon("icon.ico"))
        self.setWindowTitle("Mobile Billing Handler")
        self.setMinimumHeight(800)
        self.setMinimumWidth(800)

        # Set fonts for all the labels
        font = QFont()
        font.setFamily("Microsoft YaHei UI")
        font.setBold(True)
        for i in range(7):
            eval('self.label_' + str(i+1)).setFont(font)

        # bind buttons with their respective methods
        self.pushButton.clicked.connect(self.click_start)

    def click_start(self):
        try:
            old_filename = self.OutFilename.text()
            new_filename = self.InFilename.text()
            sheet = self.SheetName.text()
            col_phone = self.PhoneColumn.text()
            col_cost = self.CostColumn.text()
            col_count = self.CountColumn.text()
            handle_billing(old_filename, new_filename, sheet, col_phone, col_cost, col_count)
        except Exception as e:
            alert_message(str(e))
        else:
            self.resultBrowser.setText("Successfully")


def display_win():
    app = QtWidgets.QApplication(sys.argv)
    myWin = BillWin()
    myWin.show()
    sys.exit(app.exec())


def handle_billing(old_filename, new_filename, sheet, col_phone, col_cost, col_count):
    if sheet == '':
        ws_old = load_workbook(old_filename).active
    else:
        ws_old = load_workbook(old_filename)[sheet]

    phone_lst, cost_lst = read_cell(ws_old, col_phone, col_cost, col_count)

    write_cell(new_filename, phone_lst, cost_lst)


def read_cell(ws_old, col_phone, col_cost, col_count):
    phones = ws_old[col_phone]
    costs = ws_old[col_cost]
    phone_lst = []
    cost_lst = []
    if col_count == '':
        for phone, cost in zip(phones, costs):
            if phone.value is not None and len(str(phone.value)) == 11 \
                    and str(phone.value).isdigit():
                phone_lst.append(str(phone.value))
                cost_lst.append(cost.value)
    else:
        for phone, cost, count in zip(phones, costs, ws_old[col_count]):
            if phone.value is not None and len(str(phone.value)) == 11 \
                    and str(phone.value).isdigit() and count.value == '小计':
                phone_lst.append(str(phone.value))
                cost_lst.append(cost.value)
    return phone_lst, cost_lst


def write_cell(new_filename, phone_lst, cost_lst):
    wb_new = load_workbook(new_filename)
    ws_new = wb_new.active
    for phone, cost in zip(phone_lst, cost_lst):
        ws_new.append({'A': phone, 'B': cost})
    wb_new.save(new_filename)


def alert_message(message: str):
    box = QtWidgets.QMessageBox()
    box.setWindowTitle("ALERT")
    box.setText(message)
    box.addButton('Confirm', QMessageBox.YesRole)
    box.exec_()


if __name__ == '__main__':
    display_win()
